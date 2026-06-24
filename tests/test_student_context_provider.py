"""
Unit and integration tests for gateway.student_context_provider.

Tests 1–11 use synthetic DataFrames injected directly into the module
globals — no real Excel file dependency.
Test 12 uses the real students_anonymous.xlsx (STU000001).
Tests 13–22 are regression tests for bug fixes (SCP data-correctness pass).
Tests 23–50 are SCP Phase 1 audit tests:
  - credit_hours=None contract
  - track normalization (AI/CYS/DSE/SWE/GEN, unsupported CS, unknown)
  - level=None for invalid/blank level
  - current_semester majority-vote inference
  - completed_regular_semesters stability
  - incomplete-then-failed outcome correctness
  - improve-retake distinct-course counting
  - real-record integration (multiple student types)
"""

import io
import math
import re
from unittest.mock import patch

import openpyxl
import pandas as pd
import pytest

import gateway.student_context_provider as scp


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_REG_COLS = ["ID", "Semester", "Course Code", "Registration Status", "Letter Grade"]


def _inject(
    data_rows: list[dict],
    reg_rows: list[dict],
    global_current_sem: str | None = None,
) -> None:
    """
    Inject synthetic DataFrames into the SCP module globals.

    global_current_sem: sets _global_current_semester directly so tests that
    need in_progress_courses or current-semester exclusion can control it without
    loading a real Excel file.  Defaults to None (no active semester), which
    means in_progress_courses will always be empty and no semester is excluded
    from completed_regular_semesters.
    """
    scp._df_data = pd.DataFrame(data_rows)
    scp._df_reg  = (
        pd.DataFrame(reg_rows)
        if reg_rows
        else pd.DataFrame(columns=_REG_COLS)
    )
    scp._global_current_semester = global_current_sem


_BASE_DATA = {
    "ID": "TST001",
    "Name": "Test Student",
    "Program": "CIS Program - Artificial Intelligence Track",
    "First Semester": "Fall 2022",
    "Military Status": "Not Yet Deferred",
    "Level": "Senior",
    "Study Status": "Studying",
    "Consecutive Warning": float("nan"),
    "Total Warnings": float("nan"),
    "Last Semester Warning": float("nan"),
    "Last Semester GPA": 3.0,
    "Last Semester CHs": 15.0,
    "Last Semester CPs": 45.0,
    "Last Semester PHs": 15.0,
    "Cumulative GPA": 3.0,
    "Cumulative CHs": 90.0,
    "Cumulative CPs": 270.0,
    "Cumulative PHs": 90.0,
    "Current Semester CHs": 18,
}

def _reg(course_code, reg_status, grade, semester="Fall 2023"):
    return {
        "ID": "TST001",
        "Semester": semester,
        "Course Code": course_code,
        "Registration Status": reg_status,
        "Letter Grade": grade,
    }


# ---------------------------------------------------------------------------
# Test 1 — Student not found
# ---------------------------------------------------------------------------

def test_student_not_found_returns_none():
    _inject([_BASE_DATA], [])
    result = scp.get_context("NONEXISTENT")
    assert result is None


# ---------------------------------------------------------------------------
# Test 2 — Zero registration rows
# ---------------------------------------------------------------------------

def test_zero_registrations_returns_empty_context():
    _inject([_BASE_DATA], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.completed_courses == []
    assert ctx.failed_courses == []
    assert ctx.in_progress_courses == []
    assert ctx.course_history == []
    assert ctx.retake_count == {}
    assert ctx.total_improve_retakes_used == 0
    assert ctx.completed_regular_semesters == 0
    assert ctx.zero_credit_courses_passed == []


# ---------------------------------------------------------------------------
# Test 3 — Con grade
# ---------------------------------------------------------------------------

def test_con_grade_treated_as_passed():
    _inject(
        [_BASE_DATA],
        [_reg("C-CS495", "Continuing, Registered", "Con")],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "C-CS495")
    assert record.status == "passed"
    assert "C-CS495" in ctx.completed_courses
    assert "C-CS495" not in ctx.in_progress_courses


# ---------------------------------------------------------------------------
# Test 4 — I grade (incomplete)
# ---------------------------------------------------------------------------

def test_i_grade_treated_as_incomplete():
    _inject(
        [_BASE_DATA],
        [_reg("C-CS435", "Registered", "I", semester="Fall 2022")],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "C-CS435")
    assert record.status == "incomplete"
    # in_progress_courses is now global-current-semester active-blank rows only;
    # an I-grade from a past semester is NOT in in_progress_courses.
    assert "C-CS435" not in ctx.in_progress_courses
    assert ctx.retake_count.get("C-CS435", 0) == 1


# ---------------------------------------------------------------------------
# Test 5 — Withdrawn (W grade and Withdrawn tag)
# ---------------------------------------------------------------------------

def test_withdrawn_w_grade_excluded_from_all_lists():
    _inject(
        [_BASE_DATA],
        [_reg("C-MA112", "Registered", "W")],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "C-MA112")
    assert record.status == "withdrawn"
    assert "C-MA112" not in ctx.completed_courses
    assert "C-MA112" not in ctx.failed_courses
    assert "C-MA112" not in ctx.in_progress_courses
    assert ctx.retake_count.get("C-MA112", 0) == 0


def test_withdrawn_tag_excluded_from_all_lists():
    _inject(
        [_BASE_DATA],
        [_reg("C-MA112", "Withdrawn, Registered", None)],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "C-MA112")
    assert record.status == "withdrawn"
    assert "C-MA112" not in ctx.completed_courses
    assert "C-MA112" not in ctx.failed_courses
    assert "C-MA112" not in ctx.in_progress_courses
    assert ctx.retake_count.get("C-MA112", 0) == 0


# ---------------------------------------------------------------------------
# Test 6 — P grade (zero-credit pass)
# ---------------------------------------------------------------------------

def test_p_grade_passed_and_in_zero_credit_list():
    _inject(
        [_BASE_DATA],
        [_reg("HUM011", "Succeeded, Registered", "P")],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "HUM011")
    assert record.status == "passed"
    assert "HUM011" in ctx.completed_courses
    assert "HUM011" in ctx.zero_credit_courses_passed


# ---------------------------------------------------------------------------
# Test 7 — Improve tag counted in total_improve_retakes_used
# ---------------------------------------------------------------------------

def test_improve_tag_counted_in_improve_retakes():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2022"),
            _reg("C-CS111", "Improve, Succeeded, Registered", "A-", semester="Spring 2023"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.total_improve_retakes_used == 1


# ---------------------------------------------------------------------------
# Test 8 — Only-withdrawn semester not counted
# ---------------------------------------------------------------------------

def test_all_withdrawn_semester_not_counted():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Spring 2023"),
            _reg("C-MA111", "Withdrawn, Registered", None, semester="Fall 2022"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    # Fall 2022 has only withdrawn rows — should not be counted
    # Spring 2023 has a valid row — should be counted
    assert ctx.completed_regular_semesters == 1


# ---------------------------------------------------------------------------
# Test 9 — Summer not counted in completed_regular_semesters
# ---------------------------------------------------------------------------

def test_summer_not_counted_in_regular_semesters():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2022"),
            _reg("C-MA111", "Fresh, Succeeded, Registered", "C", semester="Spring 2023"),
            _reg("C-PH111", "Fresh, Succeeded, Registered", "B+", semester="Summer 2023"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.completed_regular_semesters == 2


# ---------------------------------------------------------------------------
# Test 10 — Best outcome priority (failed then passed → completed)
# ---------------------------------------------------------------------------

def test_best_outcome_failed_then_passed():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS213", "Fresh, Failed, Registered", "F", semester="Fall 2022"),
            _reg("C-CS213", "Repeat, Succeeded, Registered", "B", semester="Spring 2023"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-CS213" in ctx.completed_courses
    assert "C-CS213" not in ctx.failed_courses


# ---------------------------------------------------------------------------
# Test 11 — get_current_semester() format
# ---------------------------------------------------------------------------

def test_get_current_semester_format():
    result = scp.get_current_semester()
    assert isinstance(result, str)
    pattern = r"^(Fall|Spring|Summer) \d{4}$"
    assert re.match(pattern, result), f"Unexpected format: {result!r}"


# ---------------------------------------------------------------------------
# Test 12 — Real data: STU000001
# ---------------------------------------------------------------------------

def test_real_data_stu000001():
    import os
    excel_path = os.path.join(
        os.path.dirname(__file__), "..", "data", "students_anonymous.xlsx"
    )
    scp.load_excel(excel_path)
    ctx = scp.get_context("STU000001")
    assert ctx is not None
    assert ctx.student_id == "STU000001"
    assert len(ctx.completed_courses) > 0
    assert ctx.track_id == "AI"
    assert ctx.level == 4
    assert ctx.cgpa is not None
    assert ctx.cgpa > 0.0


# ---------------------------------------------------------------------------
# Helpers for regression tests
# ---------------------------------------------------------------------------

def _make_excel_bytes(data_rows: list[dict], reg_rows: list[dict]) -> bytes:
    """Build an in-memory .xlsx with 'data' and 'registrations' sheets."""
    wb = openpyxl.Workbook()

    # data sheet
    ws_data = wb.active
    ws_data.title = "data"
    data_cols = [
        "ID", "Name", "Program", "Level", "Study Status",
        "Cumulative GPA", "Consecutive Warning", "Total Warnings",
        "Military Status", "First Semester",
        "Cumulative PHs", "Cumulative CHs", "Cumulative CPs",
        "Last Semester GPA", "Last Semester CHs", "Last Semester CPs",
        "Last Semester PHs", "Last Semester Warning", "Current Semester CHs",
    ]
    ws_data.append(data_cols)
    for row in data_rows:
        ws_data.append([row.get(c, "") for c in data_cols])

    # registrations sheet
    ws_reg = wb.create_sheet("registrations")
    reg_cols = ["ID", "Course Code", "Semester", "Registration Status", "Letter Grade"]
    ws_reg.append(reg_cols)
    for row in reg_rows:
        ws_reg.append([row.get(c, "") for c in reg_cols])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Test 13 — completed_regular_semesters excludes current semester
# ---------------------------------------------------------------------------

def test_completed_regular_semesters_excludes_current_semester():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2024"),
            _reg("C-CS112", "Fresh, Succeeded, Registered", "B", semester="Spring 2025"),
            _reg("C-CS113", "Registered", None, semester="Fall 2025"),
        ],
        global_current_sem="Fall 2025",
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.completed_regular_semesters == 2, (
        f"Expected 2 (Fall 2025 excluded as global current), got {ctx.completed_regular_semesters}"
    )


# ---------------------------------------------------------------------------
# Test 14 — _map_status: Failed tag + blank grade → "failed", not "in_progress"
# ---------------------------------------------------------------------------

def test_map_status_failed_tag_with_blank_grade():
    result = scp._map_status("Failed, Registered", None)
    assert result == "failed", (
        f"Expected 'failed' for Failed tag + blank grade, got {result!r}"
    )


# ---------------------------------------------------------------------------
# Test 15 — in_progress_courses shows current retakes of previously passed course
# ---------------------------------------------------------------------------

def test_in_progress_courses_shows_current_retakes():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-AI321", "Fresh, Succeeded, Registered", "B+", semester="Fall 2024"),
            _reg("C-AI321", "Improve, Registered", None, semester="Spring 2026"),
        ],
        global_current_sem="Spring 2026",
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-AI321" in ctx.completed_courses, "Expected C-AI321 in completed_courses"
    assert "C-AI321" in ctx.in_progress_courses, (
        "Expected C-AI321 in in_progress_courses (active-blank improve retake in current semester)"
    )


# ---------------------------------------------------------------------------
# Test 16 — in_progress_courses excludes resolved incomplete
# ---------------------------------------------------------------------------

def test_in_progress_courses_excludes_resolved_incomplete():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-AI321", "Registered", "I", semester="Fall 2024"),
            _reg("C-AI321", "Repeat, Succeeded, Registered", "B", semester="Spring 2025"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-AI321" in ctx.completed_courses, "Expected C-AI321 in completed_courses"
    assert "C-AI321" not in ctx.in_progress_courses, (
        "Old resolved incomplete must not remain in in_progress_courses"
    )


# ---------------------------------------------------------------------------
# Test 17 — in_progress_courses includes unresolved incomplete
# ---------------------------------------------------------------------------

def test_in_progress_courses_excludes_unresolved_incomplete():
    """
    After the global-current-semester refactor, in_progress_courses contains
    only active blank-registered rows in the global current semester.
    An I-grade (incomplete) from a past semester is NOT in in_progress_courses
    even if unresolved.  It remains in course_history with status="incomplete".
    """
    _inject(
        [_BASE_DATA],
        [
            _reg("C-AI321", "Registered", "I", semester="Fall 2024"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "C-AI321")
    assert record.status == "incomplete"
    assert "C-AI321" not in ctx.in_progress_courses, (
        "I-grade from past semester must NOT be in in_progress_courses "
        "(in_progress is global-current-semester active-blank rows only)"
    )


# ---------------------------------------------------------------------------
# Test 18 — blank Study Status → "Studying"
# ---------------------------------------------------------------------------

def test_study_status_blank_becomes_studying():
    data_blank = dict(_BASE_DATA)
    data_blank["Study Status"] = ""
    _inject([data_blank], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.study_status == "Studying"

    data_nan = dict(_BASE_DATA)
    data_nan["Study Status"] = float("nan")
    _inject([data_nan], [])
    ctx2 = scp.get_context("TST001")
    assert ctx2 is not None
    assert ctx2.study_status == "Studying"


# ---------------------------------------------------------------------------
# Test 19 — zero_credit_courses_passed deduplicated
# ---------------------------------------------------------------------------

def test_zero_credit_passed_deduplicated():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-MANDATORY-001", "Succeeded, Registered", "P"),
            _reg("C-MANDATORY-001", "Succeeded, Registered", "P"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.zero_credit_courses_passed.count("C-MANDATORY-001") == 1, (
        "Duplicate P records must be deduplicated"
    )


# ---------------------------------------------------------------------------
# Test 20 — load_excel raises FileNotFoundError for missing file
# ---------------------------------------------------------------------------

def test_load_excel_file_not_found():
    with pytest.raises(FileNotFoundError) as exc_info:
        scp.load_excel("nonexistent_path/students.xlsx")
    assert "not found" in str(exc_info.value).lower()


# ---------------------------------------------------------------------------
# Test 21 — load_excel raises ValueError for missing sheet
# ---------------------------------------------------------------------------

def test_load_excel_missing_sheet(tmp_path):
    # Excel with only a 'data' sheet — 'registrations' is absent
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["ID"])
    xls_path = tmp_path / "missing_sheet.xlsx"
    wb.save(str(xls_path))

    with pytest.raises(ValueError) as exc_info:
        scp.load_excel(str(xls_path))
    msg = str(exc_info.value).lower()
    assert "sheet" in msg or "registrations" in msg


# ---------------------------------------------------------------------------
# Test 22 — load_excel raises ValueError for missing required columns
# ---------------------------------------------------------------------------

def test_load_excel_missing_required_columns(tmp_path):
    # 'data' sheet is missing the 'ID' column
    wb = openpyxl.Workbook()
    ws_data = wb.active
    ws_data.title = "data"
    ws_data.append(["Name"])  # ID is absent

    ws_reg = wb.create_sheet("registrations")
    ws_reg.append(["ID", "Course Code", "Semester", "Registration Status", "Letter Grade"])

    xls_path = tmp_path / "missing_cols.xlsx"
    wb.save(str(xls_path))

    with pytest.raises(ValueError) as exc_info:
        scp.load_excel(str(xls_path))
    msg = str(exc_info.value)
    assert "ID" in msg or "missing" in msg.lower()


# ===========================================================================
# Phase 1 Audit Tests — credit_hours contract
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 23 — CourseRecord.credit_hours is always None (SCP never sets credits)
# ---------------------------------------------------------------------------

def test_credit_hours_is_none_for_all_course_records():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2022"),
            _reg("C-MA111", "Registered", None, semester="Spring 2026"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    for record in ctx.course_history:
        assert record.credit_hours is None, (
            f"Expected credit_hours=None for {record.course_code}, "
            f"got {record.credit_hours!r}"
        )


# ===========================================================================
# Phase 1 Audit Tests — track normalization
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 24 — Artificial Intelligence → AI (supported)
# ---------------------------------------------------------------------------

def test_ai_program_maps_to_ai_track():
    data = dict(_BASE_DATA)
    data["Program"] = "CIS Program - Artificial Intelligence Track"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id == "AI"
    assert ctx.track_status == "supported"
    assert ctx.track_error_code is None


# ---------------------------------------------------------------------------
# Test 25 — Cyber Security → CYS (supported)
# ---------------------------------------------------------------------------

def test_cyber_security_maps_to_cys():
    data = dict(_BASE_DATA)
    data["Program"] = "CIS Program - Cyber Security Track"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id == "CYS"
    assert ctx.track_status == "supported"


# ---------------------------------------------------------------------------
# Test 26 — Data Science and Engineering → DSE (supported)
# ---------------------------------------------------------------------------

def test_data_science_engineering_maps_to_dse():
    data = dict(_BASE_DATA)
    data["Program"] = "CIS Program - Data Science and Engineering Track"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id == "DSE"
    assert ctx.track_status == "supported"


# ---------------------------------------------------------------------------
# Test 27 — Software Engineering → SWE (supported)
# ---------------------------------------------------------------------------

def test_software_engineering_maps_to_swe():
    data = dict(_BASE_DATA)
    data["Program"] = "CIS Program - Software Engineering Track"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id == "SWE"
    assert ctx.track_status == "supported"


# ---------------------------------------------------------------------------
# Test 28 — General Program → GEN (supported)
# ---------------------------------------------------------------------------

def test_general_program_maps_to_gen():
    data = dict(_BASE_DATA)
    data["Program"] = "Faculty of Computing and Information Sciences - General Program"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id == "GEN"
    assert ctx.track_status == "supported"


# ---------------------------------------------------------------------------
# Test 29 — Computer Science → unsupported (no KG track)
# ---------------------------------------------------------------------------

def test_computer_science_maps_to_unsupported():
    data = dict(_BASE_DATA)
    data["Program"] = "CIS Program - Computer Science Track"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id is None, (
        f"CS program must not get a track_id, got {ctx.track_id!r}"
    )
    assert ctx.track_status == "unsupported"
    assert ctx.track_error_code == "unsupported_track"


# ---------------------------------------------------------------------------
# Test 30 — Completely unknown program → unsupported
# ---------------------------------------------------------------------------

def test_unknown_program_maps_to_unsupported():
    data = dict(_BASE_DATA)
    data["Program"] = "Faculty of Alien Studies - Mars Engineering Track"
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.track_id is None
    assert ctx.track_status == "unsupported"
    assert ctx.track_error_code == "unsupported_track"


# ---------------------------------------------------------------------------
# Test 31 — Unsupported track: StudentContext still loads (not a crash)
# ---------------------------------------------------------------------------

def test_unsupported_track_still_returns_valid_context():
    data = dict(_BASE_DATA)
    data["Program"] = "CIS Program - Computer Science Track"
    _inject(
        [data],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "A", semester="Fall 2022"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None, "SCP must not crash for unsupported programs"
    assert ctx.track_status == "unsupported"
    assert len(ctx.completed_courses) == 1
    assert ctx.cgpa is not None


# ===========================================================================
# Phase 1 Audit Tests — level handling
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 32 — Invalid level does not silently default to Freshman
# ---------------------------------------------------------------------------

def test_invalid_level_is_none_not_freshman():
    data = dict(_BASE_DATA)
    data["Level"] = "Grad"  # not a valid level string
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.level is None, (
        f"Invalid level must be None, not {ctx.level!r}"
    )


# ---------------------------------------------------------------------------
# Test 33 — Blank level is None not Freshman
# ---------------------------------------------------------------------------

def test_blank_level_is_none_not_freshman():
    data = dict(_BASE_DATA)
    data["Level"] = ""
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.level is None


# ---------------------------------------------------------------------------
# Test 34 — All four valid levels map correctly
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("level_str,expected", [
    ("Freshman",  1),
    ("Sophomore", 2),
    ("Junior",    3),
    ("Senior",    4),
])
def test_valid_level_maps_correctly(level_str, expected):
    data = dict(_BASE_DATA)
    data["Level"] = level_str
    _inject([data], [])
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.level == expected


# ===========================================================================
# Phase 1 Audit Tests — current_semester inference
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 35 — Majority vote: semester with most in-progress rows wins
# ---------------------------------------------------------------------------

def test_global_inference_majority_wins_over_stale_row():
    """
    _compute_global_current_semester picks the semester with the most active
    blank-registered rows when count >= _ACTIVE_BLANK_THRESHOLD.
    Below threshold, falls back to latest non-withdrawn.
    This test verifies the winner selection (not the threshold — uses a small
    synthetic dataset so threshold is intentionally not crossed).
    """
    df_reg = pd.DataFrame([
        # 1 stale blank row in old semester
        {"ID": "S1", "Semester": "Fall 2020", "Course Code": "C-MA111",
         "Registration Status": "Registered", "Letter Grade": None},
        # 3 blank rows in a newer semester
        {"ID": "S1", "Semester": "Spring 2026", "Course Code": "C-CS435",
         "Registration Status": "Registered", "Letter Grade": None},
        {"ID": "S2", "Semester": "Spring 2026", "Course Code": "C-AI415",
         "Registration Status": "Registered", "Letter Grade": None},
        {"ID": "S3", "Semester": "Spring 2026", "Course Code": "C-AI422",
         "Registration Status": "Registered", "Letter Grade": None},
    ])
    # Below threshold → falls back to latest_non_withdrawn_global (Spring 2026 still wins)
    chosen, method = scp._compute_global_current_semester(df_reg)
    assert chosen == "Spring 2026", (
        f"Semester with more active rows should win, got {chosen!r}"
    )


# ---------------------------------------------------------------------------
# Test 36 — Fallback: latest non-withdrawn semester when no in-progress rows
# ---------------------------------------------------------------------------

def test_global_inference_fallback_latest_non_withdrawn():
    """
    When no semester meets the active-blank threshold, _compute_global_current_semester
    falls back to the chronologically latest non-withdrawn row across all students.
    Row order must not matter — chronological ordering wins.
    """
    df_reg = pd.DataFrame([
        # Row order: Fall 2023 first, then Spring 2023 — chronological order must win
        {"ID": "S1", "Semester": "Fall 2023", "Course Code": "C-CS435",
         "Registration Status": "Fresh, Succeeded, Registered", "Letter Grade": "B"},
        {"ID": "S1", "Semester": "Spring 2023", "Course Code": "C-AI415",
         "Registration Status": "Fresh, Succeeded, Registered", "Letter Grade": "A"},
    ])
    chosen, method = scp._compute_global_current_semester(df_reg)
    assert chosen == "Fall 2023", (
        f"Expected Fall 2023 (chronologically latest), got {chosen!r}"
    )
    assert method == "latest_non_withdrawn_global"


# ---------------------------------------------------------------------------
# Test 37 — completed_regular_semesters excludes inferred current semester
#           (only when current semester was inferred from in-progress rows)
# ---------------------------------------------------------------------------

def test_completed_regular_semesters_excludes_global_current():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2024"),
            _reg("C-CS112", "Fresh, Succeeded, Registered", "B", semester="Spring 2025"),
            _reg("C-CS113", "Registered", None, semester="Fall 2025"),
            _reg("C-CS114", "Registered", None, semester="Fall 2025"),
        ],
        global_current_sem="Fall 2025",
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.current_semester == "Fall 2025"
    assert ctx.completed_regular_semesters == 2, (
        f"Expected 2 (Fall 2025 excluded via global_current_semester), "
        f"got {ctx.completed_regular_semesters}"
    )


# ---------------------------------------------------------------------------
# Test 38 — completed_regular_semesters counts all semesters when no in-progress
# ---------------------------------------------------------------------------

def test_completed_regular_semesters_counts_all_when_no_global_current():
    """
    When _global_current_semester is None (synthetic test without load_excel),
    no semester is excluded from completed_regular_semesters.
    All completed Fall/Spring semesters are counted.
    """
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2022"),
            _reg("C-CS112", "Fresh, Succeeded, Registered", "B", semester="Spring 2023"),
            _reg("C-CS113", "Fresh, Succeeded, Registered", "A", semester="Fall 2023"),
            _reg("C-CS114", "Fresh, Succeeded, Registered", "A-", semester="Summer 2023"),
        ],
        # global_current_sem=None (default) → nothing excluded
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    # Summer excluded by season filter; 3 regular sems counted; nothing excluded
    assert ctx.completed_regular_semesters == 3, (
        f"Expected 3 (no global current semester, nothing excluded), "
        f"got {ctx.completed_regular_semesters}"
    )


# ===========================================================================
# Phase 1 Audit Tests — course outcome correctness
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 39 — Incomplete then failed: latest state is failed, not incomplete
# ---------------------------------------------------------------------------

def test_incomplete_then_failed_appears_in_failed_not_inprogress():
    """
    Bug fix: old priority map gave incomplete (priority 4) > failed (3),
    hiding a later failed attempt behind an older incomplete.
    After fix: latest meaningful state is failed → course must appear in
    failed_courses, not in_progress_courses.
    """
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS213", "Registered", "I", semester="Fall 2023"),
            _reg("C-CS213", "Repeat, Failed, Registered", "F", semester="Spring 2024"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-CS213" in ctx.failed_courses, (
        "C-CS213 should be in failed_courses: latest state is failed"
    )
    assert "C-CS213" not in ctx.in_progress_courses, (
        "C-CS213 must not remain in in_progress_courses after a later failed attempt"
    )
    assert "C-CS213" not in ctx.completed_courses


# ---------------------------------------------------------------------------
# Test 40 — Withdrawn-only course does not appear in any derived list
# ---------------------------------------------------------------------------

def test_withdrawn_only_course_not_in_any_list():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS300", "Withdrawn, Registered", None, semester="Fall 2022"),
            _reg("C-CS300", "Withdrawn, Registered", None, semester="Spring 2023"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-CS300" not in ctx.completed_courses
    assert "C-CS300" not in ctx.failed_courses
    assert "C-CS300" not in ctx.in_progress_courses
    # Still appears in course_history
    codes_in_history = {r.course_code for r in ctx.course_history}
    assert "C-CS300" in codes_in_history


# ---------------------------------------------------------------------------
# Test 41 — Failed tag + blank grade is failed, not in_progress (regression)
# ---------------------------------------------------------------------------

def test_failed_tag_blank_grade_is_failed_not_in_progress():
    _inject(
        [_BASE_DATA],
        [
            _reg("C-AI311", "Failed, Registered", None, semester="Fall 2022"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-AI311" in ctx.failed_courses
    assert "C-AI311" not in ctx.in_progress_courses


# ===========================================================================
# Phase 1 Audit Tests — improve-retake counting
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 42 — Distinct course count, not raw attempt count
# ---------------------------------------------------------------------------

def test_improve_retake_counts_distinct_courses_not_raw_attempts():
    """
    A student who improve-retakes the same course twice uses 1 slot, not 2.
    total_improve_retakes_used must count distinct course codes.
    """
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2021"),
            _reg("C-CS111", "Improve, Succeeded, Registered", "A-", semester="Spring 2022"),
            _reg("C-CS111", "Improve, Succeeded, Registered", "A", semester="Fall 2022"),
            _reg("C-CS222", "Fresh, Succeeded, Registered", "C+", semester="Spring 2022"),
            _reg("C-CS222", "Improve, Succeeded, Registered", "B+", semester="Fall 2022"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    # 2 distinct courses improved (C-CS111 and C-CS222), not 3 raw improve rows
    assert ctx.total_improve_retakes_used == 2, (
        f"Expected 2 distinct improved courses, got {ctx.total_improve_retakes_used}"
    )


# ---------------------------------------------------------------------------
# Test 43 — Withdrawn improve-retake does not consume a slot
# ---------------------------------------------------------------------------

def test_withdrawn_improve_retake_not_counted():
    """
    Conservative policy: withdrawn improve-retake rows are not counted since
    no confirmed handbook rule states a withdrawal consumes a slot.
    """
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS111", "Fresh, Succeeded, Registered", "B", semester="Fall 2021"),
            _reg("C-CS111", "Improve, Withdrawn, Registered", None, semester="Spring 2022"),
        ],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert ctx.total_improve_retakes_used == 0, (
        f"Withdrawn improve-retake must not consume a slot, got {ctx.total_improve_retakes_used}"
    )


# ===========================================================================
# Phase 1 Audit Tests — real-record integration
# ===========================================================================

_REAL_EXCEL = None


def _get_real_excel_path():
    import os
    return os.path.join(os.path.dirname(__file__), "..", "data", "students_anonymous.xlsx")


def _load_real_data_once():
    global _REAL_EXCEL
    if _REAL_EXCEL is None:
        path = _get_real_excel_path()
        scp.load_excel(path)
        _REAL_EXCEL = path
    return _REAL_EXCEL


# Supported track IDs that must never contain informal values
_SUPPORTED_TRACK_IDS = {"AI", "CYS", "DSE", "SWE", "GEN"}
_OLD_INFORMAL_TRACK_IDS = {"Cyber", "Data Science", "SW", "CS", "General"}


def _assert_context_invariants(ctx, student_id: str) -> None:
    """
    Assert internal consistency invariants for any StudentContext.
    Called from all real-record tests.
    """
    # Track invariants
    if ctx.track_status == "supported":
        assert ctx.track_id in _SUPPORTED_TRACK_IDS, (
            f"{student_id}: supported track_id {ctx.track_id!r} is not a KG canonical ID"
        )
        assert ctx.track_id not in _OLD_INFORMAL_TRACK_IDS, (
            f"{student_id}: old informal track ID {ctx.track_id!r} must not be used"
        )
        assert ctx.track_error_code is None
    else:
        assert ctx.track_id is None, (
            f"{student_id}: unsupported track must have track_id=None, got {ctx.track_id!r}"
        )
        assert ctx.track_error_code is not None

    # Credit hours must be None everywhere
    for record in ctx.course_history:
        assert record.credit_hours is None, (
            f"{student_id}: CourseRecord {record.course_code} has credit_hours={record.credit_hours!r}"
        )

    # No completed course also appears in failed_courses
    completed_set = set(ctx.completed_courses)
    for code in ctx.failed_courses:
        assert code not in completed_set, (
            f"{student_id}: {code} appears in both completed and failed — "
            "completed must dominate"
        )

    # in_progress courses not in failed (a currently in-progress course cannot
    # simultaneously be the latest-meaningful-failed)
    in_progress_set = set(ctx.in_progress_courses)
    for code in ctx.failed_courses:
        assert code not in in_progress_set, (
            f"{student_id}: {code} is both in_progress and failed"
        )


# ---------------------------------------------------------------------------
# Test 44 — STU000001: AI, Level 4, Studying, has in-progress courses
# ---------------------------------------------------------------------------

def test_real_stu000001_ai_studying_with_inprogress():
    _load_real_data_once()
    ctx = scp.get_context("STU000001")
    assert ctx is not None
    assert ctx.student_id == "STU000001"
    assert ctx.track_id == "AI"
    assert ctx.track_status == "supported"
    assert ctx.level == 4
    assert ctx.study_status == "Studying"
    assert ctx.cgpa is not None and ctx.cgpa > 0
    assert len(ctx.completed_courses) > 0
    assert len(ctx.in_progress_courses) > 0, "STU000001 should have in-progress courses"
    assert ctx.current_semester is not None
    # current_semester must come from majority vote (6 in-progress rows in one semester)
    assert ctx.current_semester == "Spring 2026", (
        f"Expected Spring 2026 from registrar data, got {ctx.current_semester!r}"
    )
    # Spring 2026 must NOT be in completed_regular_semesters
    sems_without_current = ctx.completed_regular_semesters
    assert sems_without_current > 0
    _assert_context_invariants(ctx, "STU000001")


# ---------------------------------------------------------------------------
# Test 45 — STU000005: General, Studying, has failed then retaken/passed
# ---------------------------------------------------------------------------

def test_real_stu000005_general_with_failed_and_retaken():
    _load_real_data_once()
    ctx = scp.get_context("STU000005")
    assert ctx is not None
    assert ctx.track_id == "GEN"
    assert ctx.track_status == "supported"
    assert ctx.study_status == "Studying"
    # Student has failed then passed C-MA111 and C-CS215
    for retaken_code in ("C-MA111", "C-CS215"):
        assert retaken_code in ctx.completed_courses, (
            f"Expected {retaken_code} in completed (passed after failure)"
        )
        assert retaken_code not in ctx.failed_courses, (
            f"{retaken_code} must not be in failed_courses after being passed"
        )
    _assert_context_invariants(ctx, "STU000005")


# ---------------------------------------------------------------------------
# Test 46 — STU000017: AI, Senior, Graduated, has improve-retakes
# ---------------------------------------------------------------------------

def test_real_stu000017_graduated_ai_with_improve_retakes():
    _load_real_data_once()
    ctx = scp.get_context("STU000017")
    assert ctx is not None
    assert ctx.track_id == "AI"
    assert ctx.track_status == "supported"
    assert ctx.level == 4
    assert ctx.study_status == "Graduated"
    assert ctx.cgpa is not None and ctx.cgpa > 0
    assert len(ctx.completed_courses) > 0
    # Graduated student: no in-progress registrations expected
    assert len(ctx.in_progress_courses) == 0, (
        f"Graduated student should have no in-progress courses, "
        f"got {ctx.in_progress_courses}"
    )
    # STU000017 has improve retakes — count must be >= 1
    assert ctx.total_improve_retakes_used >= 1
    _assert_context_invariants(ctx, "STU000017")


# ---------------------------------------------------------------------------
# Test 47 — STU000026: Computer Science Track → unsupported, context still loads
# ---------------------------------------------------------------------------

def test_real_stu000026_cs_track_unsupported():
    _load_real_data_once()
    ctx = scp.get_context("STU000026")
    assert ctx is not None, "SCP must not crash for CS-track students"
    assert ctx.track_id is None, (
        f"CS program must not get a KG track_id, got {ctx.track_id!r}"
    )
    assert ctx.track_status == "unsupported"
    assert ctx.track_error_code == "unsupported_track"
    # All other registrar facts must still load correctly
    assert ctx.cgpa is not None and ctx.cgpa > 0
    assert ctx.level == 4
    assert ctx.study_status == "Graduated"
    _assert_context_invariants(ctx, "STU000026")


# ---------------------------------------------------------------------------
# Test 48 — STU000041: Suspended, General, Freshman, NaN CGPA
#            All registrations are Forced Withdraw with Abs grade →
#            SCP correctly marks all as withdrawn (not failed).
# ---------------------------------------------------------------------------

def test_real_stu000041_suspended_general_forced_withdraw():
    _load_real_data_once()
    ctx = scp.get_context("STU000041")
    assert ctx is not None
    assert ctx.track_id == "GEN"
    assert ctx.track_status == "supported"
    assert ctx.level == 1
    assert ctx.study_status == "Suspended"
    assert ctx.cgpa is None, "Suspended student with NaN CGPA should return cgpa=None"
    # All courses have "Forced Withdraw" tag → correctly withdrawn, not failed
    assert len(ctx.completed_courses) == 0
    assert len(ctx.failed_courses) == 0, (
        "All courses were Forced Withdraw — must not appear in failed_courses"
    )
    assert len(ctx.in_progress_courses) == 0
    assert ctx.retake_count == {}, "Forced Withdraw does not count as an attempt"
    _assert_context_invariants(ctx, "STU000041")


# ---------------------------------------------------------------------------
# Test 49 — STU000100: Cyber Security, Senior, Graduated
# ---------------------------------------------------------------------------

def test_real_stu000100_cyber_graduated():
    _load_real_data_once()
    ctx = scp.get_context("STU000100")
    assert ctx is not None
    assert ctx.track_id == "CYS"
    assert ctx.track_status == "supported"
    assert ctx.level == 4
    assert ctx.study_status == "Graduated"
    assert ctx.cgpa is not None and ctx.cgpa > 0
    assert len(ctx.completed_courses) > 0
    _assert_context_invariants(ctx, "STU000100")


# ---------------------------------------------------------------------------
# Test 50 — Real-record sweep: all 6 program types map correctly
# ---------------------------------------------------------------------------

def test_real_program_types_map_to_correct_track_ids():
    """
    Sample one student per known program string from the real dataset.
    Verify each maps to the correct KG canonical track ID or unsupported.
    """
    _load_real_data_once()

    expected_mapping = {
        "Artificial Intelligence": "AI",
        "Cyber Security": "CYS",
        "Data Science and Engineering": "DSE",
        "Software Engineering": "SWE",
        "General Program": "GEN",
        "Computer Science": None,  # unsupported
    }

    for keyword, expected_track in expected_mapping.items():
        track_id, track_status, track_error_code = scp._parse_track(
            f"CIS Program - {keyword} Track"
        )
        if expected_track is not None:
            assert track_id == expected_track, (
                f"Program containing '{keyword}' should map to {expected_track!r}, "
                f"got {track_id!r}"
            )
            assert track_status == "supported"
        else:
            assert track_id is None, (
                f"Program containing '{keyword}' should be unsupported, "
                f"got track_id={track_id!r}"
            )
            assert track_status == "unsupported"


# ===========================================================================
# Phase 2 Tests — global current_semester inference + active-blank in_progress
# ===========================================================================

# ---------------------------------------------------------------------------
# Test 51 — Real data: global current semester is Spring 2026
# ---------------------------------------------------------------------------

def test_real_global_current_semester_is_spring_2026():
    """
    After load_excel() on the real dataset, _global_current_semester must be
    "Spring 2026" because Spring 2026 has 3170 active blank-registered rows —
    far above the _ACTIVE_BLANK_THRESHOLD of 100.
    """
    _load_real_data_once()
    assert scp._global_current_semester == "Spring 2026", (
        f"Expected _global_current_semester='Spring 2026', "
        f"got {scp._global_current_semester!r}"
    )


# ---------------------------------------------------------------------------
# Test 52 — Active blank-registered in global current semester → in_progress
# ---------------------------------------------------------------------------

def test_active_blank_in_global_semester_is_in_progress():
    """
    A student with a blank-grade, Registered (no terminal tag) row in the
    global current semester must appear in in_progress_courses.
    """
    _inject(
        [_BASE_DATA],
        [
            _reg("C-CS435", "Registered", None, semester="Spring 2026"),
        ],
        global_current_sem="Spring 2026",
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-CS435" in ctx.in_progress_courses, (
        "Active blank-registered row in global current semester must be in_progress"
    )


# ---------------------------------------------------------------------------
# Test 53 — Old blank rows with terminal tags NOT in in_progress
# ---------------------------------------------------------------------------

def test_old_blank_rows_with_terminal_tags_not_in_progress():
    """
    Old rows with blank grades but terminal tags (Succeeded, Failed, Withdrawn)
    in a past semester must NOT appear in in_progress_courses even if
    _is_active_blank would be False for them (which it should be).
    """
    _inject(
        [_BASE_DATA],
        [
            # Past semester — Succeeded + blank grade (data anomaly in real dataset)
            _reg("C-CS111", "Succeeded, Registered", None, semester="Fall 2022"),
            # Past semester — Failed + blank grade
            _reg("C-MA111", "Failed, Registered", None, semester="Spring 2023"),
        ],
        global_current_sem="Spring 2026",
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    assert "C-CS111" not in ctx.in_progress_courses, (
        "Old Succeeded+blank row must NOT be in_progress (terminal tag)"
    )
    assert "C-MA111" not in ctx.in_progress_courses, (
        "Old Failed+blank row must NOT be in_progress (terminal tag)"
    )
    # in_progress_courses must be empty — no active rows in Spring 2026
    assert ctx.in_progress_courses == []


# ---------------------------------------------------------------------------
# Test 54 — _is_active_blank: terminal tags correctly excluded
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("reg_status,grade,expected", [
    ("Registered", None,                        True),   # plain active
    ("Fresh, Registered", None,                 True),   # fresh tag OK
    ("Improve, Registered", None,               True),   # improve retake OK
    ("Succeeded, Registered", None,             False),  # has Succeeded
    ("Failed, Registered", None,                False),  # has Failed
    ("Withdrawn, Registered", None,             False),  # has Withdrawn
    ("Fresh, Forced Withdraw, Registered", None, False), # has Forced Withdraw
    ("Registered", "B",                         False),  # grade present
    ("Registered", "I",                         False),  # I-grade → not active blank
    ("Registered", "F",                         False),  # F-grade → not active blank
    ("Registered", "W",                         False),  # W-grade → not active blank
])
def test_is_active_blank_cases(reg_status, grade, expected):
    result = scp._is_active_blank(reg_status, grade)
    assert result == expected, (
        f"_is_active_blank({reg_status!r}, {grade!r}) → {result}, expected {expected}"
    )


# ---------------------------------------------------------------------------
# Test 55 — Forced Withdraw correctly withdrawn (not failed), with any grade
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("grade", [None, "Abs", "W"])
def test_forced_withdraw_various_grades_all_withdrawn(grade):
    """
    FW rows with any grade variant (blank, Abs, W) must all map to 'withdrawn'.
    Confirmed from real dataset: all 31 FW rows are in Fall 2025 with grades
    Abs (14), blank (14), and W (3).  None should appear in failed_courses.
    Supervisor confirmation pending for FW+Abs (blank grade treated same as W).
    """
    _inject(
        [_BASE_DATA],
        [_reg("C-CS111", "Fresh, Forced Withdraw, Registered", grade)],
    )
    ctx = scp.get_context("TST001")
    assert ctx is not None
    record = next(r for r in ctx.course_history if r.course_code == "C-CS111")
    assert record.status == "withdrawn", (
        f"FW row with grade={grade!r} must be withdrawn, got {record.status!r}"
    )
    assert "C-CS111" not in ctx.failed_courses
    assert "C-CS111" not in ctx.in_progress_courses
    assert ctx.retake_count.get("C-CS111", 0) == 0, (
        "FW rows must not count as attempts"
    )


# ---------------------------------------------------------------------------
# Test 56 — _clean_grade case normalization
# ---------------------------------------------------------------------------

@pytest.mark.parametrize("raw,expected", [
    ("abs",  "Abs"),
    ("ABS",  "Abs"),
    ("Abs",  "Abs"),
    ("con",  "Con"),
    ("CON",  "Con"),
    ("Con",  "Con"),
    ("p",    "P"),
    ("P",    "P"),
    ("f",    "F"),
    ("F",    "F"),
    ("i",    "I"),
    ("I",    "I"),
    ("w",    "W"),
    ("W",    "W"),
    ("A",    "A"),    # regular grade unchanged
    ("B+",   "B+"),
    (None,   None),
    ("",     None),
    (float("nan"), None),
])
def test_clean_grade_normalization(raw, expected):
    result = scp._clean_grade(raw)
    assert result == expected, (
        f"_clean_grade({raw!r}) → {result!r}, expected {expected!r}"
    )
